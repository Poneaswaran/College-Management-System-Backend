"""
GraphQL queries for grades management
"""
import strawberry
from typing import List, Optional
from strawberry.types import Info
import base64
from io import BytesIO

from grades.models import CourseGrade, SemesterGPA, StudentCGPA
from profile_management.models import StudentProfile, Semester
from .types import CourseGradeType, SemesterGPAType, StudentCGPAType
from core.graphql.auth import require_auth


# ==================================================
# ADDITIONAL TYPES
# ==================================================

@strawberry.type
class GradeOverviewType:
    """Overall grade statistics"""
    cgpa: float
    cgpa_status: str  # "Excellent", "Very Good", "Good", etc.
    total_credits: float
    credits_earned: float
    performance_trend: str
    semester_gpas: List[SemesterGPAType]


@strawberry.type
class ExportGradesType:
    """Excel export response"""
    success: bool
    message: str
    file_base64: Optional[str]
    filename: str


@strawberry.type
class GradesQuery:
    """Grades-related queries"""
    
    @strawberry.field
    @require_auth
    def my_grades(
        self,
        info: Info,
        register_number: str,
        semester_id: Optional[int] = None
    ) -> List[CourseGradeType]:
        """
        Get all course grades for a student
        Optionally filter by semester
        """
        try:
            student = StudentProfile.objects.get(register_number=register_number)
            
            query = CourseGrade.objects.filter(
                student=student,
                is_published=True
            ).select_related(
                'subject',
                'subject__department',
                'semester',
                'semester__academic_year'
            ).order_by('-semester__start_date', 'subject__code')
            
            if semester_id:
                query = query.filter(semester_id=semester_id)
            
            return list(query)
            
        except StudentProfile.DoesNotExist:
            return []
    
    @strawberry.field
    @require_auth
    def grade_overview(
        self,
        info: Info,
        register_number: str
    ) -> Optional[GradeOverviewType]:
        """
        Get overall grade overview including CGPA and semester GPAs
        """
        try:
            student = StudentProfile.objects.get(register_number=register_number)
            
            # Get or calculate CGPA
            cgpa_record = StudentCGPA.calculate_cgpa(student)
            
            if not cgpa_record:
                return None
            
            # Determine CGPA status
            cgpa_value = float(cgpa_record.cgpa)
            if cgpa_value >= 9.0:
                cgpa_status = "Excellent"
            elif cgpa_value >= 8.0:
                cgpa_status = "Very Good"
            elif cgpa_value >= 7.0:
                cgpa_status = "Good"
            elif cgpa_value >= 6.0:
                cgpa_status = "Satisfactory"
            else:
                cgpa_status = "Needs Improvement"
            
            # Get all semester GPAs
            semester_gpas = SemesterGPA.objects.filter(
                student=student
            ).select_related(
                'semester',
                'semester__academic_year'
            ).order_by('-semester__start_date')
            
            return GradeOverviewType(
                cgpa=float(cgpa_record.cgpa),
                cgpa_status=cgpa_status,
                total_credits=float(cgpa_record.total_credits),
                credits_earned=float(cgpa_record.credits_earned),
                performance_trend=cgpa_record.performance_trend,
                semester_gpas=list(semester_gpas)
            )
            
        except StudentProfile.DoesNotExist:
            return None
    
    @strawberry.field
    @require_auth
    def semester_grades(
        self,
        info: Info,
        register_number: str,
        semester_id: int
    ) -> List[CourseGradeType]:
        """Get grades for a specific semester"""
        try:
            student = StudentProfile.objects.get(register_number=register_number)
            
            grades = CourseGrade.objects.filter(
                student=student,
                semester_id=semester_id,
                is_published=True
            ).select_related(
                'subject',
                'subject__department',
                'semester',
                'semester__academic_year'
            ).order_by('subject__code')
            
            return list(grades)
            
        except StudentProfile.DoesNotExist:
            return []
    
    @strawberry.field
    @require_auth
    def export_grades(
        self,
        info: Info,
        register_number: str,
        semester_id: Optional[int] = None
    ) -> ExportGradesType:
        """
        Export grades to Excel format
        Returns base64 encoded Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            student = StudentProfile.objects.get(register_number=register_number)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Grade Report"
            
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:I1')
            ws['A1'] = f"Academic Grade Report - {student.full_name}"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Student Info
            ws['A2'] = f"Register Number: {student.register_number}"
            ws['A3'] = f"Department: {student.department.name if student.department else 'N/A'}"
            ws['A4'] = f"Course: {student.course.name if student.course else 'N/A'}"
            
            # Get CGPA
            cgpa_record = StudentCGPA.objects.filter(student=student).first()
            if cgpa_record:
                ws['E2'] = f"Overall CGPA: {cgpa_record.cgpa}"
                ws['E3'] = f"Total Credits: {cgpa_record.total_credits}"
                ws['E4'] = f"Credits Earned: {cgpa_record.credits_earned}"
            
            # Headers
            headers = ['Course Code', 'Course Name', 'Credits', 'Marks', 'Grade', 'Grade Points', 'Semester', 'Exam Date', 'Exam Type']
            ws.append([])  # Empty row
            ws.append(headers)
            
            header_row = ws.max_row
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Get grades
            query = CourseGrade.objects.filter(
                student=student,
                is_published=True
            ).select_related('subject', 'semester', 'semester__academic_year')
            
            if semester_id:
                query = query.filter(semester_id=semester_id)
            
            grades = query.order_by('-semester__start_date', 'subject__code')
            
            # Add data
            for grade in grades:
                semester_name = f"{grade.semester.get_number_display()} {grade.semester.academic_year.year_code}"
                exam_date_str = grade.exam_date.strftime('%Y-%m-%d') if grade.exam_date else 'N/A'
                
                row_data = [
                    grade.subject.code,
                    grade.subject.name,
                    float(grade.credits),
                    f"{float(grade.total_marks)}/{float(grade.total_max_marks)} ({float(grade.percentage):.1f}%)",
                    grade.grade,
                    float(grade.grade_points),
                    semester_name,
                    exam_date_str,
                    grade.exam_type
                ]
                ws.append(row_data)
                
                # Style data cells
                current_row = ws.max_row
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_num != 2 else 'left')
            
            # Adjust column widths
            column_widths = [15, 35, 10, 20, 10, 12, 20, 15, 15]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
            
            # Add semester-wise GPA summary
            semester_gpas = SemesterGPA.objects.filter(student=student).order_by('-semester__start_date')
            if semester_gpas.exists():
                ws.append([])  # Empty row
                ws.append(['Semester-wise GPA Summary'])
                ws.append(['Semester', 'GPA', 'Credits', 'Credits Earned'])
                
                for sem_gpa in semester_gpas:
                    sem_name = f"{sem_gpa.semester.get_number_display()} {sem_gpa.semester.academic_year.year_code}"
                    ws.append([
                        sem_name,
                        float(sem_gpa.gpa),
                        float(sem_gpa.total_credits),
                        float(sem_gpa.credits_earned)
                    ])
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Encode to base64
            file_base64 = base64.b64encode(excel_file.getvalue()).decode('utf-8')
            
            filename = f"grades_{student.register_number}_{semester_id if semester_id else 'all'}.xlsx"
            
            return ExportGradesType(
                success=True,
                message="Grade report generated successfully",
                file_base64=file_base64,
                filename=filename
            )
            
        except StudentProfile.DoesNotExist:
            return ExportGradesType(
                success=False,
                message="Student not found",
                file_base64=None,
                filename=""
            )
        except Exception as e:
            return ExportGradesType(
                success=False,
                message=f"Error generating report: {str(e)}",
                file_base64=None,
                filename=""
            )
