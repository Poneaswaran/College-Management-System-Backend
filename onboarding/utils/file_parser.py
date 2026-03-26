import csv
import hashlib
import os
from io import TextIOWrapper
from itertools import islice

from openpyxl import load_workbook


def compute_file_hash(file_obj):
    hasher = hashlib.sha256()
    for chunk in file_obj.chunks():
        hasher.update(chunk)
    file_obj.seek(0)
    return hasher.hexdigest()


def get_extension(filename):
    return os.path.splitext(filename)[1].lower()


def parse_csv(file_obj):
    file_obj.seek(0)
    wrapper = TextIOWrapper(file_obj.file, encoding="utf-8-sig")
    reader = csv.DictReader(wrapper)
    rows = [normalize_row(row) for row in reader]
    file_obj.seek(0)
    return rows


def parse_xlsx(file_obj):
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = []
    headers = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if index == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue

        row_dict = {}
        for i, value in enumerate(row):
            header = headers[i] if i < len(headers) else f"col_{i}"
            row_dict[header] = value
        rows.append(normalize_row(row_dict))

    file_obj.seek(0)
    return rows


def normalize_row(row):
    normalized = {}
    for key, value in row.items():
        clean_key = str(key).strip() if key is not None else ""
        if isinstance(value, str):
            normalized[clean_key] = value.strip()
        else:
            normalized[clean_key] = value
    return normalized


def parse_upload_file(file_obj, extension):
    if extension == ".csv":
        return parse_csv(file_obj)
    if extension == ".xlsx":
        return parse_xlsx(file_obj)
    raise ValueError("Unsupported file format")


def iter_csv_rows(file_obj):
    file_obj.seek(0)
    wrapper = TextIOWrapper(file_obj.file, encoding="utf-8-sig")
    reader = csv.DictReader(wrapper)
    headers = [str(h).strip() for h in (reader.fieldnames or [])]
    try:
        for row in reader:
            yield normalize_row(row), headers
    finally:
        file_obj.seek(0)


def iter_xlsx_rows(file_obj):
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    worksheet = workbook.active

    headers = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if index == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue

        row_dict = {}
        for i, value in enumerate(row):
            header = headers[i] if i < len(headers) else f"col_{i}"
            row_dict[header] = value
        yield normalize_row(row_dict), headers

    file_obj.seek(0)


def iter_upload_rows(file_obj, extension):
    if extension == ".csv":
        return iter_csv_rows(file_obj)
    if extension == ".xlsx":
        return iter_xlsx_rows(file_obj)
    raise ValueError("Unsupported file format")


def iter_chunked_rows(row_iterator, chunk_size):
    while True:
        chunk = list(islice(row_iterator, chunk_size))
        if not chunk:
            break
        yield chunk
